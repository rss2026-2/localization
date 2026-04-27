#!/usr/bin/env python3

import os
from pathlib import Path

from rosbag2_py import SequentialReader, StorageOptions, ConverterOptions
from rclpy.serialization import deserialize_message
from rosidl_runtime_py.utilities import get_message

import matplotlib.pyplot as plt
from openpyxl import Workbook


def read_timing_topic(bag_path, topic_name):
    reader = SequentialReader()

    storage_options = StorageOptions(
        uri=str(bag_path),
        storage_id="sqlite3"
    )
    converter_options = ConverterOptions(
        input_serialization_format="cdr",
        output_serialization_format="cdr"
    )

    reader.open(storage_options, converter_options)

    # Get topic types
    topic_types = reader.get_all_topics_and_types()
    type_map = {t.name: t.type for t in topic_types}

    if topic_name not in type_map:
        print(f"⚠️ Topic {topic_name} not found in {bag_path}")
        return []

    msg_type = get_message(type_map[topic_name])

    values = []

    while reader.has_next():
        topic, data, t = reader.read_next()

        if topic == topic_name:
            msg = deserialize_message(data, msg_type)
            values.append(msg.data  * 1e-9)

    return values


def pad_lists(data_dict):
    max_len = max(len(v) for v in data_dict.values())

    padded = {}
    for key, values in data_dict.items():
        padded[key] = values + [None] * (max_len - len(values))

    return padded


def write_sheet(ws, data_dict):
    # Write headers
    for col_idx, run_name in enumerate(data_dict.keys(), start=1):
        ws.cell(row=1, column=col_idx, value=run_name)

    # Write data
    for col_idx, (run_name, values) in enumerate(data_dict.items(), start=1):
        for row_idx, val in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=val)


def main():
    bag_paths = [
        "timing_simple_hall_all_topics_no_noise_1_d",
         "timing_simple_hall_all_topics_no_noise_2_d",
         "timing_simple_hall_all_topics_no_noise_3_d",
         "timing_simple_hall_all_topics_no_noise_4_d",
     ]

    motion_data = {}
    sensor_data = {}

    for bag in bag_paths:
        bag_path = Path(bag)
        run_name = bag_path.name

        print(f"📂 Processing {run_name}")

        motion = read_timing_topic(bag_path, "/timing/motion_model")
        sensor = read_timing_topic(bag_path, "/timing/sensor_model")

        motion_data[run_name] = motion
        sensor_data[run_name] = sensor

    # Pad so all columns align
    motion_data = pad_lists(motion_data)
    sensor_data = pad_lists(sensor_data)

    # Create Excel workbook
    wb_motion = Workbook()

    # Motion sheet
    ws_motion = wb_motion.active
    ws_motion.title = "motion_model"
    # ws_motion = wb_motion.create_sheet(title="sensor_model")
    write_sheet(ws_motion, motion_data)

    wb_sensor = Workbook()
    # Sensor sheet
    ws_sensor = wb_sensor.active
    ws_sensor.title = "sensor_model"
    # ws_sensor = wb_sensor.create_sheet(title="sensor_model")
    write_sheet(ws_sensor, sensor_data)

    # Save file
    output_file_motion = "timing_results_motion.xlsx"
    wb_motion.save(output_file_motion)
    print(f"✅ Saved to {output_file_motion}")

    output_file_sensor = "timing_results_sensor_new_timing_method_d.xlsx"
    wb_sensor.save(output_file_sensor)
    print(f"✅ Saved to {output_file_sensor}")


    plot_boxplot(motion_data, "Motion Model Timing", "motion_boxplot.png")
    plot_boxplot(sensor_data, "Sensor Model Timing", "sensor_boxplot.png")

def plot_boxplot(data_dict, title, output_file):
    plt.figure(figsize=(10, 6))

    labels = list(data_dict.keys())

    # Remove None padding before plotting
    data = [
        [v for v in values if v is not None]
        for values in data_dict.values()
    ]

    plt.boxplot(data, labels=labels)

    plt.xlabel("Trials")
    plt.ylabel("Time (seconds)")
    plt.title(title)
    plt.grid(True)

    plt.tight_layout()
    plt.savefig(output_file)
    plt.close()

if __name__ == "__main__":
    main()
